"""build_progress.py — emit student-progress.json for the booking page's
"returning student" lookup (card-per-lesson view).

Keyed by SHA-256(lowercased email). Value = per-lesson status array so the page
can show one card per lesson (done / upcoming / open) and offer a reschedule or
book button on the ones that aren't done yet. The file carries NO name, email or
contact — only lesson numbers, statuses and (for upcoming) the booked date/time.
Worst case if public: someone who already knows a student's exact email sees that
student's lesson statuses. They cannot enumerate or download the roster.

Completion is inferred by DATE (a lesson dated before today = already taken),
because the Schedule sheet only carries Scheduled/Rescheduled statuses.
"""
import sys, json, hashlib, datetime as dt
sys.stdout.reconfigure(encoding="utf-8")
from pathlib import Path
from zoneinfo import ZoneInfo
import openpyxl

LON = ZoneInfo("Europe/London")
XLSX = Path(r"G:/My Drive/AI_Development/02_freelance/03_ai-teaching/students/attendance.xlsx")
OUT = Path(__file__).parent / "docs" / "student-progress.json"
TOTAL = 5
DEAD = {"cancelled", "canceled", "dropped", "void"}

def h(email):
    return hashlib.sha256(email.strip().lower().encode()).hexdigest()

def main():
    today = dt.datetime.now(LON).date()
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    email_of = {}
    for r in wb["Students"].iter_rows(min_row=4, values_only=True):   # header row 3
        if r and r[0] and len(r) > 1 and r[1] and "@" in str(r[1]):
            # Skip dropped/cancelled students — they must NOT be lookup-able or
            # re-bookable on the public page, even though their past lessons stay
            # in the sheet as a record. (Vince Yiu dropped out 2026-07-17.)
            sstatus = str(r[2]).strip().lower() if len(r) > 2 and r[2] else ""
            if sstatus in DEAD:
                continue
            email_of[str(r[0]).strip()] = str(r[1]).strip().lower()
    # per student -> per lesson -> best row
    per = {}   # name -> {lesson: {"date":d, "uk":t}}
    for r in wb["Schedule"].iter_rows(min_row=5, values_only=True):   # header row 4
        if not r or not r[0] or not r[1]:
            continue
        status = str(r[4]).strip().lower() if len(r) > 4 and r[4] else ""
        if status in DEAD:
            continue
        try:
            lesson = int(r[1])
        except (TypeError, ValueError):
            continue
        if not (1 <= lesson <= TOTAL):
            continue
        name = str(r[0]).strip()
        datestr = str(r[2])[:10] if len(r) > 2 and r[2] else ""
        uk = str(r[3])[:5] if len(r) > 3 and r[3] else ""
        try:
            d = dt.date.fromisoformat(datestr) if datestr else None
        except ValueError:
            d = None
        rec = per.setdefault(name, {})
        # keep the most relevant row per lesson: prefer the latest date
        cur = rec.get(lesson)
        if cur is None or (d and cur.get("_d") and d > cur["_d"]) or (d and not cur.get("_d")):
            rec[lesson] = {"_d": d, "date": datestr, "uk": uk}

    out = {}
    for name, rec in per.items():
        email = email_of.get(name)
        if not email:
            continue
        lessons = []
        for n in range(1, TOTAL + 1):
            row = rec.get(n)
            if not row or not row.get("_d"):
                lessons.append({"n": n, "status": "open"})
            elif row["_d"] < today:
                lessons.append({"n": n, "status": "done"})
            else:
                lessons.append({"n": n, "status": "upcoming",
                                "date": row["date"], "uk": row["uk"]})
        done = sum(1 for l in lessons if l["status"] == "done")
        out[h(email)] = {"done": done, "total": TOTAL, "lessons": lessons}

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "generated": dt.datetime.now(LON).isoformat(timespec="minutes"),
        "note": "keyed by sha256(lowercased email); per-lesson statuses only, no PII",
        "students": out,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✅ wrote progress for {len(out)} students (hashed) -> {OUT}")

if __name__ == "__main__":
    main()
